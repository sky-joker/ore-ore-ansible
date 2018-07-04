#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.colors import WHITE
import re
import json
import argparse

def options():
    parser = argparse.ArgumentParser(prog="ansible-result2excel.py",
                                     add_help=True,
                                     description="Tools to convert ansible execution results (JSON) to Excel")

    parser.add_argument("--file", "-f",
                        type=str, required=True,
                        help="Specify JSON file to convert")

    parser.add_argument("--output", "-o",
                        type=str, default="output.xlsx",
                        help="Specify the Excel file name to output(default: output.xlsx)")

    args = parser.parse_args()
    return args

def fill_config(type, color):
    fill = PatternFill(fill_type=type,
                       fgColor=color)

    return fill

def border_config(style, color):
    border = Border(
        left=Side(
            border_style=style,
            color=color
        ),
        right=Side(
            border_style=style,
            color=color
        ),
        top=Side(
            border_style=style,
            color=color
        ),
        bottom=Side(
            border_style=style,
            color=color
        ),
    )

    return border

def main():
    args = options()

    wb = Workbook()
    ws = wb.active

    ws.title = "Ansible Result"

    ws.append([])
    ws.append(["", "host", "ok", "changed", "unreachable", "failed", "skipped"])

    for cell in ws[2]:
        if(getattr(cell, "coordinate") != "A2"):
            cell.border = border_config("thin", "30A5FF")
            cell.fill = fill_config("solid", "30A5FF")
            cell.font = Font(color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    with open(args.file, "r") as f:
        j = json.loads(f.read())

        for host_name in j['stats'].keys():
            tmp = [""]
            tmp.append(host_name)
            tmp.append(j["stats"][host_name]["ok"])
            tmp.append(j["stats"][host_name]["changed"])
            tmp.append(j["stats"][host_name]["unreachable"])
            tmp.append(j["stats"][host_name]["failures"])
            tmp.append(j["stats"][host_name]["skipped"])
            ws.append(tmp)
            for cell in list(ws).pop():
                if(not(re.match(r"A[0-9]+", getattr(cell, "coordinate")))):
                    cell.border = border_config("thin", "30A5FF")

                if(cell.col_idx - 1 == 5 and cell.value > 0):
                    cell.fill = fill_config("solid", "FC4549")

        check = {}
        for play in j['plays']:
            for task in play['tasks']:
                ws.append([])
                task_name = task["task"]["name"]
                task_id = task["task"]["id"]

                if(not(task_id in check)):
                    for key in task['hosts'].keys():
                        tmp = [""]
                        tmp.append('task name')
                        tmp.append('host')
                        for key2 in task['hosts'][key].keys():
                            tmp.append(key2)
                    ws.append(tmp)
                else:
                    check[task_id] = True

                for cell in list(ws).pop():
                    if(not(re.match(r"A[0-9]+", getattr(cell, "coordinate")))):
                        cell.border = border_config("thin", "000000")
                        cell.fill = fill_config("solid", "81FF88")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                for key in task['hosts'].keys():
                    tmp = [""]
                    tmp.append(task_name)
                    tmp.append(key)
                    for key2 in task['hosts'][key].keys():
                        tmp.append(str(task['hosts'][key][key2]))

                    ws.append(tmp)

                    for cell in list(ws).pop():
                        if(not(re.match(r"A[0-9]+", getattr(cell, "coordinate")))):
                            cell.border = border_config("thin", "000000")
                    """
                    if(cell.value == "True"):
                        cell.fill = fill_config("solid", "7FBDFF")

                    if(cell.value == "False"):
                        cell.fill = fill_config("solid", "FC4549")
                    """

    wb.save(args.output)

if __name__ == "__main__":
    main()
